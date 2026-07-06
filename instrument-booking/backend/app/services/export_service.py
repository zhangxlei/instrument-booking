import uuid
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.booking import Booking
from app.models.instrument import Instrument
from app.models.user import User


async def export_bookings_excel(
    db: AsyncSession,
    start_date: str | None = None,
    end_date: str | None = None,
    instrument_id: str | None = None,
    status_filter: str | None = None,
) -> Workbook:
    query = select(Booking).order_by(Booking.created_at.desc())

    if start_date:
        query = query.where(Booking.created_at >= datetime.fromisoformat(start_date))
    if end_date:
        query = query.where(Booking.created_at <= datetime.fromisoformat(end_date))
    if instrument_id:
        query = query.where(Booking.instrument_id == uuid.UUID(instrument_id))
    if status_filter:
        query = query.where(Booking.status == status_filter)

    result = await db.execute(query)
    bookings = result.scalars().all()

    user_ids = {b.user_id for b in bookings}
    inst_ids = {b.instrument_id for b in bookings}
    user_map = {}
    if user_ids:
        users = await db.execute(select(User).where(User.id.in_(user_ids)))
        user_map = {u.id: u for u in users.scalars().all()}
    inst_map = {}
    if inst_ids:
        insts = await db.execute(select(Instrument).where(Instrument.id.in_(inst_ids)))
        inst_map = {i.id: i for i in insts.scalars().all()}

    status_labels = {
        "pending": "待审批", "approved": "已通过", "rejected": "已拒绝",
        "cancelled": "已取消", "completed": "已完成",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "预约记录"
    ws.append(["预约人", "仪器名称", "开始时间", "结束时间", "状态", "目的", "备注", "捎话", "拒绝原因", "创建时间"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for b in bookings:
        u = user_map.get(b.user_id)
        inst = inst_map.get(b.instrument_id)
        ws.append([
            f"{u.full_name}({u.username})" if u else str(b.user_id),
            inst.name if inst else str(b.instrument_id),
            b.start_time.strftime("%Y-%m-%d %H:%M"),
            b.end_time.strftime("%Y-%m-%d %H:%M"),
            status_labels.get(b.status, b.status),
            b.purpose or "", b.notes or "", b.message or "", b.rejection_reason or "",
            b.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    return wb


async def export_instruments_excel(db: AsyncSession) -> Workbook:
    result = await db.execute(select(Instrument).order_by(Instrument.name))
    instruments = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "仪器列表"
    ws.append(["名称", "位置", "状态", "需要审批", "创建时间"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for i in instruments:
        ws.append([
            i.name, i.location or "", i.status,
            "是" if i.requires_approval else "否",
            i.created_at.strftime("%Y-%m-%d %H:%M"),
        ])

    return wb
